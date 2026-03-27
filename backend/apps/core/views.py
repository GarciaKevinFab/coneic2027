import io
import logging
from collections import OrderedDict
from decimal import Decimal

from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import generics, permissions, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.certificates.models import Certificate
from apps.participants.models import Participant, ParticipantType
from apps.payments.models import PaymentLog
from apps.tickets.models import Ticket, TicketType
from apps.users.permissions import IsOrganizer
from apps.workshops.models import Workshop, WorkshopEnrollment

from .serializers import (
    AdminParticipantSerializer,
    AdminPaymentSerializer,
    AdminStatsSerializer,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Pagination
# ---------------------------------------------------------------------------
class AdminPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 500


# ---------------------------------------------------------------------------
# Admin Participants
# ---------------------------------------------------------------------------
class AdminParticipantsView(generics.ListAPIView):
    """
    GET /api/v1/admin/participants/
    List all participants with rich filtering for the admin dashboard.

    Query params:
    - search: filter by name or email (partial match)
    - type: filter by participant type name (student/professional/speaker/organizer)
    - payment_status: filter by payment status (pending/paid/cancelled)
    - is_accredited: filter by accreditation (true/false)
    - university: filter by university (partial match)
    - ticket_type: filter by ticket type ID
    - ordering: sort field (default: -created_at)
    """

    serializer_class = AdminParticipantSerializer
    permission_classes = [permissions.IsAdminUser | IsOrganizer]
    pagination_class = AdminPagination

    def get_queryset(self):
        queryset = (
            Participant.objects.select_related(
                "user",
                "participant_type",
                "ticket__ticket_type",
                "accredited_by",
            )
            .prefetch_related("selected_workshops")
            .all()
        )

        # Search by name or email
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(user__full_name__icontains=search)
                | Q(user__email__icontains=search)
            )

        # Filter by participant type
        p_type = self.request.query_params.get("type")
        if p_type:
            queryset = queryset.filter(participant_type__name=p_type)

        # Filter by payment status
        payment_status = self.request.query_params.get("payment_status")
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)

        # Filter by accreditation
        is_accredited = self.request.query_params.get("is_accredited")
        if is_accredited is not None:
            queryset = queryset.filter(
                is_accredited=is_accredited.lower() in ("true", "1", "yes")
            )

        # Filter by university
        university = self.request.query_params.get("university")
        if university:
            queryset = queryset.filter(user__university__icontains=university)

        # Filter by ticket type
        ticket_type = self.request.query_params.get("ticket_type")
        if ticket_type:
            queryset = queryset.filter(ticket__ticket_type_id=ticket_type)

        # Ordering
        ordering = self.request.query_params.get("ordering", "-created_at")
        allowed_orderings = {
            "created_at",
            "-created_at",
            "user__full_name",
            "-user__full_name",
            "payment_status",
            "-payment_status",
        }
        if ordering in allowed_orderings:
            queryset = queryset.order_by(ordering)

        return queryset


# ---------------------------------------------------------------------------
# Export Participants to Excel
# ---------------------------------------------------------------------------
class ExportParticipantsView(APIView):
    """
    GET /api/v1/admin/participants/export/
    Export all participants to an Excel file (.xlsx) using openpyxl.
    Admin/Organizer only.
    """

    permission_classes = [permissions.IsAdminUser | IsOrganizer]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return Response(
                {"detail": "openpyxl no esta instalado en el servidor."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        participants = (
            Participant.objects.select_related(
                "user",
                "participant_type",
                "ticket__ticket_type",
            )
            .prefetch_related("selected_workshops")
            .order_by("user__full_name")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participantes CONEIC 2027"

        # -- Header row -------------------------------------------------------
        headers = [
            "ID",
            "Nombre Completo",
            "Email",
            "Telefono",
            "Universidad",
            "Carrera",
            "Ciudad",
            "Pais",
            "Tipo de Participante",
            "Estado de Pago",
            "Acreditado",
            "Fecha de Acreditacion",
            "Tipo de Entrada",
            "Monto Pagado",
            "Talleres Inscritos",
            "Fecha de Registro",
        ]

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="1A237E", end_color="1A237E", fill_type="solid"
        )
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # -- Data rows --------------------------------------------------------
        for row_idx, p in enumerate(participants, start=2):
            workshops = ", ".join(
                w.name for w in p.selected_workshops.all()
            )
            ticket_type = (
                p.ticket.ticket_type.name if p.ticket and p.ticket.ticket_type else ""
            )
            amount = str(p.ticket.amount_paid) if p.ticket else ""

            row_data = [
                p.id,
                p.user.full_name,
                p.user.email,
                p.user.phone,
                p.user.university,
                p.user.career,
                p.user.city,
                p.user.country,
                p.participant_type.get_name_display() if p.participant_type else "",
                p.get_payment_status_display(),
                "Si" if p.is_accredited else "No",
                (
                    p.accredited_at.strftime("%Y-%m-%d %H:%M")
                    if p.accredited_at
                    else ""
                ),
                ticket_type,
                amount,
                workshops,
                p.created_at.strftime("%Y-%m-%d %H:%M"),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # -- Column widths ----------------------------------------------------
        column_widths = [6, 30, 30, 15, 35, 25, 15, 12, 18, 15, 12, 20, 18, 12, 40, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width

        # -- Auto-filter ------------------------------------------------------
        ws.auto_filter.ref = ws.dimensions

        # -- Freeze top row ---------------------------------------------------
        ws.freeze_panes = "A2"

        # -- Write to response ------------------------------------------------
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"participantes_coneic2027_{timestamp}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        logger.info(
            "Participants exported to Excel by %s (%d records)",
            request.user.email,
            participants.count(),
        )
        return response


# ---------------------------------------------------------------------------
# Admin Payments
# ---------------------------------------------------------------------------
class AdminPaymentsView(generics.ListAPIView):
    """
    GET /api/v1/admin/payments/
    List all payment logs with filters for the admin dashboard.

    Query params:
    - status: filter by payment status (pending/success/failed/refunded)
    - payment_method: filter by method (culqi/yape)
    - search: search by charge ID, user name, or email
    - ordering: sort field (default: -created_at)
    """

    serializer_class = AdminPaymentSerializer
    permission_classes = [permissions.IsAdminUser | IsOrganizer]
    pagination_class = AdminPagination

    def get_queryset(self):
        queryset = PaymentLog.objects.select_related(
            "ticket__user",
            "ticket__ticket_type",
        ).all()

        # Filter by status
        pay_status = self.request.query_params.get("status")
        if pay_status:
            queryset = queryset.filter(status=pay_status)

        # Filter by payment method
        method = self.request.query_params.get("payment_method")
        if method:
            queryset = queryset.filter(payment_method=method)

        # Search
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(culqi_charge_id__icontains=search)
                | Q(ticket__user__full_name__icontains=search)
                | Q(ticket__user__email__icontains=search)
                | Q(ticket__purchase_code__icontains=search)
            )

        # Ordering
        ordering = self.request.query_params.get("ordering", "-created_at")
        allowed = {"created_at", "-created_at", "amount", "-amount", "status"}
        if ordering in allowed:
            queryset = queryset.order_by(ordering)

        return queryset


# ---------------------------------------------------------------------------
# Admin Accreditation (by QR token)
# ---------------------------------------------------------------------------
class AdminAccreditView(APIView):
    """
    POST /api/v1/admin/accredit/
    Accredit a participant by their QR token.
    Body: { "qr_token": "<uuid>" }
    """

    permission_classes = [permissions.IsAdminUser | IsOrganizer]

    def post(self, request):
        qr_token = request.data.get("qr_token")
        if not qr_token:
            return Response(
                {"detail": "El campo qr_token es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            participant = Participant.objects.select_related(
                "user", "participant_type", "ticket__ticket_type"
            ).get(qr_token=qr_token)
        except Participant.DoesNotExist:
            return Response(
                {"detail": "No se encontro un participante con este codigo QR."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if participant.is_accredited:
            return Response(
                {
                    "detail": "Este participante ya fue acreditado.",
                    "participant": AdminParticipantSerializer(participant).data,
                    "already_accredited": True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        if participant.payment_status != Participant.PaymentStatus.PAID:
            return Response(
                {
                    "detail": "El participante no tiene un pago confirmado.",
                    "participant": AdminParticipantSerializer(participant).data,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        participant.is_accredited = True
        participant.accredited_at = timezone.now()
        participant.accredited_by = request.user
        participant.save(
            update_fields=["is_accredited", "accredited_at", "accredited_by"]
        )

        logger.info(
            "Participant %s accredited by %s",
            participant.user.email,
            request.user.email,
        )

        return Response(
            {
                "detail": "Participante acreditado exitosamente.",
                "participant": AdminParticipantSerializer(participant).data,
            },
            status=status.HTTP_200_OK,
        )


# ---------------------------------------------------------------------------
# Admin Dashboard Stats
# ---------------------------------------------------------------------------
class AdminStatsView(APIView):
    """
    GET /api/v1/admin/stats/
    Returns aggregate statistics for the admin dashboard.
    """

    permission_classes = [permissions.IsAdminUser | IsOrganizer]

    def get(self, request):
        # -- Participants -----------------------------------------------------
        total_participants = Participant.objects.count()
        accredited = Participant.objects.filter(is_accredited=True).count()

        payment_counts = dict(
            Participant.objects.values_list("payment_status")
            .annotate(count=Count("id"))
            .values_list("payment_status", "count")
        )

        participants_by_type = list(
            Participant.objects.values("participant_type__name")
            .annotate(count=Count("id"))
            .order_by("participant_type__name")
        )

        # -- Revenue ----------------------------------------------------------
        total_revenue = (
            PaymentLog.objects.filter(status=PaymentLog.Status.SUCCESS).aggregate(
                total=Sum("amount")
            )["total"]
            or Decimal("0.00")
        )

        # -- Tickets ----------------------------------------------------------
        tickets_sold = Ticket.objects.filter(status=Ticket.Status.CONFIRMED).count()
        tickets_by_type = list(
            Ticket.objects.filter(status=Ticket.Status.CONFIRMED)
            .values("ticket_type__name")
            .annotate(count=Count("id"))
            .order_by("ticket_type__name")
        )

        # -- Workshops --------------------------------------------------------
        total_workshops = Workshop.objects.filter(is_active=True).count()
        total_enrollments = WorkshopEnrollment.objects.count()

        # -- Certificates -----------------------------------------------------
        certificates_issued = Certificate.objects.count()

        stats = {
            "total_participants": total_participants,
            "accredited_count": accredited,
            "pending_accreditation": total_participants - accredited,
            "payment_paid": payment_counts.get("paid", 0),
            "payment_pending": payment_counts.get("pending", 0),
            "payment_cancelled": payment_counts.get("cancelled", 0),
            "total_revenue": total_revenue,
            "tickets_sold": tickets_sold,
            "tickets_by_type": tickets_by_type,
            "total_workshops": total_workshops,
            "total_enrollments": total_enrollments,
            "participants_by_type": participants_by_type,
            "certificates_issued": certificates_issued,
        }

        serializer = AdminStatsSerializer(stats)
        return Response(serializer.data, status=status.HTTP_200_OK)


# ---------------------------------------------------------------------------
# Admin Workshops Report
# ---------------------------------------------------------------------------
class AdminWorkshopsReportView(APIView):
    """
    GET /api/v1/admin/workshops/report/
    Returns a summary report of all workshops with enrollment stats.
    """

    permission_classes = [permissions.IsAdminUser | IsOrganizer]

    def get(self, request):
        workshops = (
            Workshop.objects.select_related("speaker")
            .annotate(
                actual_enrolled=Count("enrollments"),
            )
            .order_by("start_time")
        )

        report = []
        for ws in workshops:
            report.append(
                {
                    "id": ws.id,
                    "name": ws.name,
                    "workshop_type": ws.workshop_type,
                    "workshop_type_display": ws.get_workshop_type_display(),
                    "speaker": ws.speaker.name if ws.speaker else None,
                    "start_time": ws.start_time.isoformat() if ws.start_time else None,
                    "end_time": ws.end_time.isoformat() if ws.end_time else None,
                    "location": ws.location,
                    "capacity": ws.capacity,
                    "enrolled_count": ws.enrolled_count,
                    "actual_enrolled": ws.actual_enrolled,
                    "available_slots": ws.available_slots,
                    "is_full": ws.is_full,
                    "occupancy_percent": (
                        round((ws.enrolled_count / ws.capacity) * 100, 1)
                        if ws.capacity > 0
                        else 0
                    ),
                }
            )

        return Response(
            {
                "total_workshops": len(report),
                "workshops": report,
            },
            status=status.HTTP_200_OK,
        )
